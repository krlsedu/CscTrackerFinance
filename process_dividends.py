import os
import glob
import datetime
import uuid
import openpyxl


def to_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip()
        # Remove "R$" and spaces
        clean = val.replace("R$", "").replace(" ", "")
        if "," in clean:
            clean = clean.replace(".", "").replace(",", ".")
        try:
            return float(clean)
        except ValueError:
            return 0.0
    return 0.0


def to_currency_str(val):
    if val is None:
        return "R$ 0,00"
    if isinstance(val, (int, float)):
        parts = f"{val:.2f}".split(".")
        integer_part = parts[0]
        decimal_part = parts[1]

        # Add thousands separator for Portuguese format
        reversed_integer = integer_part[::-1]
        groups = [
            reversed_integer[i : i + 3] for i in range(0, len(reversed_integer), 3)
        ]
        formatted_integer = ".".join(groups)[::-1]

        return f"R$ {formatted_integer},{decimal_part}"
    if isinstance(val, str):
        val = val.strip()
        if not val.startswith("R$"):
            # Check if it looks like a number and format it
            f_val = to_float(val)
            return to_currency_str(f_val)
        return val
    return str(val)


def format_quantity(val):
    if val is None:
        return "0"
    if isinstance(val, (int, float)):
        if val == int(val):
            return str(int(val))
        return str(val)
    if isinstance(val, str):
        val = val.strip()
        try:
            val_float = float(val.replace(",", "."))
            if val_float == int(val_float):
                return str(int(val_float))
            return str(val_float)
        except ValueError:
            return val
    return str(val)


def format_date(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                dt = datetime.datetime.strptime(val, fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue
    return str(val)


def escape_sql(val):
    if val is None:
        return ""
    # Escape single quotes by doubling them
    return str(val).replace("'", "''")


def main():
    excel_files = glob.glob(os.path.join("dividendos", "*.xlsx"))
    print(f"Encontrados {len(excel_files)} arquivos excel na pasta dividendos.")

    all_transactions = []

    for file_path in excel_files:
        print(f"Processando arquivo: {file_path}")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            max_r = sheet.max_row

            # Iterate starting from row 2 (skipping header)
            for r in range(2, max_r + 1):
                # Columns:
                # 1: Produto
                # 2: Pagamento
                # 3: Tipo de Evento
                # 4: Instituição
                # 5: Quantidade
                # 6: Preço unitário
                # 7: Valor líquido
                produto = sheet.cell(row=r, column=1).value
                pagamento = sheet.cell(row=r, column=2).value
                tipo_evento = sheet.cell(row=r, column=3).value
                instituicao = sheet.cell(row=r, column=4).value
                quantidade = sheet.cell(row=r, column=5).value
                preco_unitario = sheet.cell(row=r, column=6).value
                valor_liquido = sheet.cell(row=r, column=7).value

                # Filter out empty/total/summary rows
                if produto is None:
                    continue
                produto_str = str(produto).strip()
                if (
                    not produto_str
                    or produto_str == ""
                    or produto_str.lower() == "total"
                ):
                    continue

                # Extrai o name antes do primeiro hífen
                # Exemplo: BCRI11 - BANESTES RECEBIVEIS IMOBILIARIOS FII -> BCRI11
                if "-" in produto_str:
                    name = produto_str.split("-")[0].strip()
                else:
                    name = produto_str

                date_formatted = format_date(pagamento)
                if not date_formatted:
                    # Skip rows with invalid or missing date
                    continue

                # Converte e formata os valores numéricos para o campo de texto
                qty_str = format_quantity(quantidade)
                pu_str = to_currency_str(preco_unitario)
                vl_str = to_currency_str(valor_liquido)
                vl_float = to_float(valor_liquido)

                # Monta o text conforme a especificação:
                # "Tipo de evento" + recebido, referente a "Quantidade" cotas de "Produto" no valor de "Preço unitário" por cota, total: "Valor líquido" na instituição "Instituiçao"
                tipo_evento_str = (
                    str(tipo_evento).strip() if tipo_evento is not None else ""
                )
                instituicao_str = (
                    str(instituicao).strip() if instituicao is not None else ""
                )

                text = f"{tipo_evento_str} recebido, referente a {qty_str} cotas de {produto_str} no valor de {pu_str} por cota, total: {vl_str} na instituição {instituicao_str}"

                # Cria chave uuid única para key
                key_uuid = str(uuid.uuid4())

                # Cria uuid para request_id seguindo o padrão CscTrackerBff-uuid
                req_uuid = f"CscTrackerBff-{uuid.uuid4()}"

                # Cria data e hora atual para last_update
                now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")

                # Monta o dicionário com os dados limpos e escapados para SQL
                transaction = {
                    "date": date_formatted,
                    "type": "income",
                    "value": vl_float,
                    "name": escape_sql(name),
                    "package_name": "null",
                    "app_name": "B3",
                    "text": escape_sql(text),
                    "user_id": 1,
                    "last_update": now_str,
                    "category": "Proventos",
                    "key": key_uuid,
                    "copy": "null",
                    "request_id": f"'{req_uuid}'",
                    "is_installment": "N",
                    "installment_id": "null",
                }

                all_transactions.append(transaction)

        except Exception as e:
            print(f"Erro ao processar arquivo {file_path}: {e}")

    # Ordena as transações por data de forma ascendente
    all_transactions.sort(key=lambda t: t["date"])

    # Gera os scripts INSERT SQL
    sql_lines = []
    for t in all_transactions:
        sql = (
            f"INSERT INTO public.transactions (date, type, value, name, package_name, app_name, text, user_id, "
            f"last_update, category, key, copy, request_id, is_installment, installment_id) "
            f"VALUES ('{t['date']}', '{t['type']}', {t['value']}, '{t['name']}', {t['package_name']}, "
            f"'{t['app_name']}', '{t['text']}', {t['user_id']}, '{t['last_update']}', '{t['category']}', "
            f"'{t['key']}', {t['copy']}, {t['request_id']}, '{t['is_installment']}', {t['installment_id']});"
        )
        sql_lines.append(sql)

    output_path = "dividendos.sql"
    with open(output_path, "w", encoding="utf-8") as f_out:
        for line in sql_lines:
            f_out.write(line + "\n")

    print(
        f"\nSucesso! {len(sql_lines)} inserts SQL salvos com sucesso em '{output_path}'."
    )


if __name__ == "__main__":
    main()
