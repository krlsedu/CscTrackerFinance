import unittest
from unittest.mock import MagicMock
from service.TransactionHandler import TransactionHandler
from csctracker_py_core.utils.utils import Utils


class TestTransactionHandler(unittest.TestCase):
    def setUp(self):
        self.remote_repository = MagicMock()
        self.http_repository = MagicMock()
        self.http_repository.get_headers.return_value = {"Authorization": "Bearer test-token"}
        self.remote_repository.get_objects.return_value = []  # No existing transactions, to avoid 'Ignored'
        Utils.inform_to_client = MagicMock()
        self.handler = TransactionHandler(self.remote_repository, self.http_repository)

    def test_devolucao_iof_cashback(self):
        json_info = {
            "packageName": "com.test.app",
            "appName": "TestApp",
            "postTime": "1773752400000",  # some epoch timestamp
            "key": "test_key"
        }
        text_str = "Devolvemos o IOF de R$ 0,15 referente à compra na Paypal *Jetbrainsam. Continue aproveitando o cartão com o menor custo do mercado."
        
        self.handler.transaction(text_str, json_info)
        
        # Verify that insert was called
        self.remote_repository.insert.assert_called_once()
        inserted_data = self.remote_repository.insert.call_args[1]["data"]
        
        # Assertions
        self.assertEqual(inserted_data["type"], "income")
        self.assertEqual(inserted_data["value"], 0.15)
        self.assertEqual(inserted_data.get("category"), "Cashback")
        self.assertEqual(inserted_data["name"], "IOF TestApp")
        self.assertEqual(inserted_data["app_name"], "TestApp")
        self.assertEqual(inserted_data["text"], text_str)

    def test_standard_purchase(self):
        json_info = {
            "packageName": "com.test.app",
            "appName": "TestApp",
            "postTime": "1773752400000",
            "key": "test_key"
        }
        text_str = "Compra de R$ 50,00 em Loja Teste às 15:30."
        
        self.handler.transaction(text_str, json_info)
        
        self.remote_repository.insert.assert_called_once()
        inserted_data = self.remote_repository.insert.call_args[1]["data"]
        
        self.assertEqual(inserted_data["type"], "outcome")
        self.assertEqual(inserted_data["value"], 50.00)
        self.assertIsNone(inserted_data.get("category"))
        self.assertEqual(inserted_data["name"], "Loja Teste")

    def test_nubank_cashback_outcome(self):
        json_info = {
            "packageName": "com.nubank",
            "appName": "Nubank",
            "postTime": "1773752400000",
            "key": "nubank_key"
        }
        text_str = "Compra de R$ 100,00 em Loja Teste às 15:30."
        
        self.handler.transaction(text_str, json_info)
        
        # Verify that insert was called twice (once for original, once for cashback)
        self.assertEqual(self.remote_repository.insert.call_count, 2)
        
        # Check first insert (original transaction)
        first_call_args = self.remote_repository.insert.call_args_list[0]
        inserted_data_1 = first_call_args[1]["data"]
        self.assertEqual(inserted_data_1["type"], "outcome")
        self.assertEqual(inserted_data_1["value"], 100.00)
        self.assertEqual(inserted_data_1["app_name"], "Nubank")
        self.assertIsNone(inserted_data_1.get("category"))
        self.assertEqual(inserted_data_1["text"], text_str)
        
        # Check second insert (cashback transaction)
        second_call_args = self.remote_repository.insert.call_args_list[1]
        inserted_data_2 = second_call_args[1]["data"]
        self.assertEqual(inserted_data_2["type"], "income")
        self.assertEqual(inserted_data_2["name"], "Nubank")
        self.assertEqual(inserted_data_2["app_name"], "Nubank")
        self.assertEqual(inserted_data_2["value"], 1.25)
        self.assertEqual(inserted_data_2["text"], f"Cashback de 1.25 referente a {text_str}")
        self.assertEqual(inserted_data_2["key"], "nubank_key_cashback")

    def test_nubank_cashback_ignored(self):
        # Setup get_objects to return some existing transaction, making it ignored
        self.remote_repository.get_objects.return_value = [{"id": 123}]
        
        json_info = {
            "packageName": "com.nubank",
            "appName": "Nubank",
            "postTime": "1773752400000",
            "key": "nubank_key"
        }
        text_str = "Compra de R$ 100,00 em Loja Teste às 15:30."
        
        self.handler.transaction(text_str, json_info)
        
        # Verify that insert was called only once (because the cashback was not generated since category is Ignored)
        self.assertEqual(self.remote_repository.insert.call_count, 1)
        inserted_data = self.remote_repository.insert.call_args[1]["data"]
        self.assertEqual(inserted_data["category"], "Ignored")

    def test_nubank_cashback_not_new_but_no_cashback_exists(self):
        # We call save_transaction directly with a transaction that has an 'id'
        # Under the new logic, since no cashback exists, it should still generate cashback!
        transaction = {
            "id": "existing-id",
            "type": "outcome",
            "value": 100.00,
            "app_name": "Nubank",
            "text": "Compra de R$ 100,00 em Loja Teste às 15:30.",
            "key": "nubank_key",
            "date": "2026-07-16"
        }
        
        # Make get_objects return empty list for cashback check
        self.remote_repository.get_objects.return_value = []
        
        self.handler.save_transaction(transaction)
        
        # Verify that insert was called twice (once for original, once for cashback)
        self.assertEqual(self.remote_repository.insert.call_count, 2)

    def test_nubank_cashback_already_exists_not_ignored(self):
        # If cashback already exists and is not 'Ignored', it should NOT generate another cashback
        transaction = {
            "type": "outcome",
            "value": 100.00,
            "app_name": "Nubank",
            "text": "Compra de R$ 100,00 em Loja Teste às 15:30.",
            "key": "nubank_key",
            "date": "2026-07-16"
        }
        
        # Mock get_objects to return an existing cashback transaction that is not Ignored
        def mock_get_objects(table, keys=None, data=None, headers=None):
            if data and data.get('key') == "nubank_key_cashback":
                return [{"id": "cashback-id", "category": "Cashback"}]
            return []
            
        self.remote_repository.get_objects.side_effect = mock_get_objects
        
        self.handler.save_transaction(transaction)
        
        # Verify that insert was called only once (no cashback generated)
        self.assertEqual(self.remote_repository.insert.call_count, 1)

    def test_nubank_cashback_already_exists_but_ignored(self):
        # If cashback already exists but is 'Ignored', it SHOULD generate another cashback
        transaction = {
            "type": "outcome",
            "value": 100.00,
            "app_name": "Nubank",
            "text": "Compra de R$ 100,00 em Loja Teste às 15:30.",
            "key": "nubank_key",
            "date": "2026-07-16"
        }
        
        # Mock get_objects to return an existing cashback transaction that is Ignored
        def mock_get_objects(table, keys=None, data=None, headers=None):
            if data and data.get('key') == "nubank_key_cashback":
                return [{"id": "cashback-id", "category": "Ignored"}]
            return []
            
        self.remote_repository.get_objects.side_effect = mock_get_objects
        
        self.handler.save_transaction(transaction)
        
        # Verify that insert was called twice (original + new cashback)
        self.assertEqual(self.remote_repository.insert.call_count, 2)

    def test_nubank_cashback_installments(self):
        json_info = {
            "packageName": "com.nubank",
            "appName": "Nubank",
            "postTime": "1773752400000",
            "key": "nubank_key"
        }
        # A transaction with (3x) installments
        text_str = "Compra de R$ 300,00 em Loja Teste às 15:30 (3x)."
        
        self.handler.transaction(text_str, json_info)
        
        # Verify that insert was called 4 times:
        # - 3 times for the original split transactions
        # - 1 time for the full cashback transaction (triggered only on the first installment)
        self.assertEqual(self.remote_repository.insert.call_count, 4)
        
        # Let's verify the first installment and its cashback
        # First insert: installment 1 (mutated reference or final loop text depending on implementation, but let's check its value)
        inserted_1 = self.remote_repository.insert.call_args_list[0][1]["data"]
        self.assertEqual(inserted_1["type"], "outcome")
        self.assertEqual(inserted_1["value"], 100.00)
        
        # Second insert: cashback for installment 1 (copied, so preserves the installment text and gets full cashback value)
        inserted_2 = self.remote_repository.insert.call_args_list[1][1]["data"]
        self.assertEqual(inserted_2["type"], "income")
        self.assertEqual(inserted_2["name"], "Nubank")
        self.assertEqual(inserted_2["value"], 3.75)  # 1.25% of 300.00
        self.assertTrue("1/3" in inserted_2["text"])
        self.assertEqual(inserted_2["text"], f"Cashback de 3.75 referente a Compra de R$ 300,00 em Loja Teste às 15:30 (3x). 1/3")

    def test_nubank_cashback_installment_edit(self):
        # Scenario: We are editing an existing installment transaction of Nubank.
        # It has 3 installments in the database.
        # We call save_transactions with an updated installment.
        
        # Mock get_objects to return the 3 existing installments when queried by save_transactions
        existing_installments = [
            {
                "id": "inst-1",
                "key": "nubank_key",
                "is_installment": "S",
                "installment_id": "inst-group-123",
                "text": "Compra de R$ 300,00 em Loja Teste às 15:30 (3x). 1/3",
                "value": 100.00,
                "app_name": "Nubank",
                "type": "outcome",
                "category": "OldCategory",
                "name": "OldName",
                "date": "2026-07-16"
            },
            {
                "id": "inst-2",
                "key": "nubank_key",
                "is_installment": "S",
                "installment_id": "inst-group-123",
                "text": "Compra de R$ 300,00 em Loja Teste às 15:30 (3x). 2/3",
                "value": 100.00,
                "app_name": "Nubank",
                "type": "outcome",
                "category": "OldCategory",
                "name": "OldName",
                "date": "2026-08-16"
            },
            {
                "id": "inst-3",
                "key": "nubank_key",
                "is_installment": "S",
                "installment_id": "inst-group-123",
                "text": "Compra de R$ 300,00 em Loja Teste às 15:30 (3x). 3/3",
                "value": 100.00,
                "app_name": "Nubank",
                "type": "outcome",
                "category": "OldCategory",
                "name": "OldName",
                "date": "2026-09-16"
            }
        ]
        
        # Setup mocks:
        # First call: save_transactions calls get_objects to get existing installments.
        # Second call: check_and_save_cashback calls get_objects to see if cashback exists. We return empty list (no cashback exists).
        def mock_get_objects(table, keys=None, data=None, headers=None):
            if data and data.get('installment_id') == "inst-group-123":
                return existing_installments
            if data and data.get('key') == "nubank_key_cashback":
                return []
            return []
            
        self.remote_repository.get_objects.side_effect = mock_get_objects
        
        # This is the updated transaction sent to save_transactions
        edited_transaction = {
            "key": "nubank_key",
            "is_installment": "S",
            "installment_id": "inst-group-123",
            "text": "Compra de R$ 300,00 em Loja Teste às 15:30 (3x). 1/3",
            "category": "NewCategory",
            "name": "NewName"
        }
        
        self.handler.save_transactions([edited_transaction], headers={"Authorization": "Bearer test-token"})
        
        # Verify that self.remote_repository.insert was called for:
        # - The 3 updated installments (inserts/upserts them with new category/name)
        # - The new cashback transaction generated for the first installment
        # Total insert calls should be 4.
        self.assertEqual(self.remote_repository.insert.call_count, 4)
        
        # Let's verify that the cashback transaction was generated correctly
        cashback_call = None
        for call in self.remote_repository.insert.call_args_list:
            data = call[1]["data"]
            if data.get("type") == "income" and "cashback" in data.get("key", ""):
                cashback_call = call
                break
                
        self.assertIsNotNone(cashback_call, "Cashback insert call not found!")
        cashback_data = cashback_call[1]["data"]
        self.assertEqual(cashback_data["type"], "income")
        self.assertEqual(cashback_data["name"], "Nubank")
        self.assertEqual(cashback_data["category"], "Cashback")
        self.assertEqual(cashback_data["value"], 3.75)  # 1.25% of (100.00 * 3)
        self.assertTrue("1/3" in cashback_data["text"])

    def test_process_b3_dividends(self):
        import io
        import openpyxl
        from unittest.mock import MagicMock
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Produto", "Pagamento", "Tipo de Evento", "Instituição", "Quantidade", "Preço unitário", "Valor líquido"])
        ws.append(["BCRI11 - BANESTES RECEBIVEIS", "15/07/2026", "Rendimento", "BANCO BTG PACTUAL S/A.", 7, "R$ 0,79", "R$ 5,53"])
        ws.append(["BRCR11 - FDO INV IMOB", "15/07/2026", "Rendimento", "NU INVESTIMENTOS S.A. - CTVM", 22, "R$ 0,41", "R$ 9,02"])
        ws.append(["TSMC34 - TAIWAN SEMICONDUCTOR", "15/07/2026", "Dividendo", "BANCO INTER S.A.", 4, "R$ 0,46", "R$ 1,82"])
        ws.append(["TOTAL", "", "", "", "", "", ""])
        
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        headers = {"Authorization": "Bearer test-token", "userName": "test@test.com"}
        self.remote_repository.get_objects.return_value = []
        self.remote_repository.get_user.return_value = {"id": 1, "email": "test@test.com"}
        
        file_mock = MagicMock()
        file_mock.read.return_value = file_stream.getvalue()
        
        result = self.handler.process_b3_dividends(file_mock, headers)
        
        self.assertEqual(result['status'], 'success')
        self.assertEqual(result['processed'], 3)
        self.assertEqual(result['inserted'], 3)
        
        # 3 calls to get_objects to check dividends_b3 table
        self.assertEqual(self.remote_repository.get_objects.call_count, 3)
        # 6 calls to insert (3 for transactions, 3 for dividends_b3)
        self.assertEqual(self.remote_repository.insert.call_count, 6)
        
        # Verify first transaction call
        first_tx_call = self.remote_repository.insert.call_args_list[0][1]["data"]
        self.assertEqual(first_tx_call["name"], "BCRI11")
        self.assertEqual(first_tx_call["app_name"], "BTG")
        self.assertEqual(first_tx_call["value"], 5.53)
        self.assertEqual(first_tx_call["type"], "income")
        self.assertEqual(first_tx_call["category"], "Proventos")
        self.assertEqual(first_tx_call["user_id"], 1)
        
        # Verify first dividends_b3 call
        first_div_call = self.remote_repository.insert.call_args_list[1][1]["data"]
        self.assertEqual(first_div_call["ticker"], "BCRI11")
        self.assertEqual(first_div_call["data_pagamento"], "2026-07-15")
        self.assertEqual(first_div_call["tipo_evento"], "Rendimento")
        self.assertEqual(first_div_call["quantidade"], 7)
        self.assertEqual(first_div_call["preco_unitario"], 0.79)
        self.assertEqual(first_div_call["user_id"], 1)

        # Verify second transaction (NU INVESTIMENTOS -> Nubank)
        second_tx_call = self.remote_repository.insert.call_args_list[2][1]["data"]
        self.assertEqual(second_tx_call["name"], "BRCR11")
        self.assertEqual(second_tx_call["app_name"], "Nubank")
        self.assertEqual(second_tx_call["user_id"], 1)
        
        # Verify third transaction (BANCO INTER -> BANCO INTER)
        third_tx_call = self.remote_repository.insert.call_args_list[4][1]["data"]
        self.assertEqual(third_tx_call["name"], "TSMC34")
        self.assertEqual(third_tx_call["app_name"], "BANCO INTER S.A.")
        self.assertEqual(third_tx_call["user_id"], 1)

    def test_process_b3_dividends_with_duplicates(self):
        import io
        import openpyxl
        from unittest.mock import MagicMock
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Produto", "Pagamento", "Tipo de Evento", "Instituição", "Quantidade", "Preço unitário", "Valor líquido"])
        ws.append(["BCRI11 - BANESTES RECEBIVEIS", "15/07/2026", "Rendimento", "BANCO BTG PACTUAL S/A.", 7, "R$ 0,79", "R$ 5,53"])
        
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        headers = {"Authorization": "Bearer test-token", "userName": "test@test.com"}
        
        # get_objects returns an existing item, meaning it is a duplicate!
        self.remote_repository.get_objects.return_value = [{"id": 123}]
        self.remote_repository.get_user.return_value = {"id": 1, "email": "test@test.com"}
        
        file_mock = MagicMock()
        file_mock.read.return_value = file_stream.getvalue()
        
        result = self.handler.process_b3_dividends(file_mock, headers)
        
        self.assertEqual(result['status'], 'success')
        self.assertEqual(result['processed'], 1)
        self.assertEqual(result['inserted'], 0) # No insertion
        
        # Verify we still called get_objects
        self.assertEqual(self.remote_repository.get_objects.call_count, 1)
        # But we did not call insert
        self.remote_repository.insert.assert_not_called()

    def test_process_b3_dividends_with_user_and_request_id(self):
        import io
        import openpyxl
        from unittest.mock import MagicMock, patch
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Produto", "Pagamento", "Tipo de Evento", "Instituição", "Quantidade", "Preço unitário", "Valor líquido"])
        ws.append(["BCRI11 - BANESTES RECEBIVEIS", "15/07/2026", "Rendimento", "BANCO BTG PACTUAL S/A.", 7, "R$ 0,79", "R$ 5,53"])
        
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        headers = {"Authorization": "Bearer test-token", "userName": "john.doe@test.com"}
        self.remote_repository.get_objects.return_value = []
        self.remote_repository.get_user.return_value = {"id": 99, "email": "john.doe@test.com"}
        
        file_mock = MagicMock()
        file_mock.read.return_value = file_stream.getvalue()
        
        with patch('csctracker_py_core.utils.request_info.RequestInfo.get_correlation_id', return_value="my-custom-req-123"):
            result = self.handler.process_b3_dividends(file_mock, headers)
            
        self.assertEqual(result['status'], 'success')
        self.assertEqual(result['processed'], 1)
        self.assertEqual(result['inserted'], 1)
        
        # Verify transaction insert data
        tx_call_data = self.remote_repository.insert.call_args_list[0][1]["data"]
        self.assertEqual(tx_call_data["user_id"], 99)
        self.assertEqual(tx_call_data["request_id"], "my-custom-req-123")
        
        # Verify dividends_b3 insert data
        div_call_data = self.remote_repository.insert.call_args_list[1][1]["data"]
        self.assertEqual(div_call_data["user_id"], 99)
        self.assertEqual(div_call_data["request_id"], "my-custom-req-123")
        self.assertIsNotNone(div_call_data["last_update"])
        self.assertIsNotNone(div_call_data["created_at"])

    def _build_ofx(self):
        return (
            "<OFX>\n"
            "<BANKTRANLIST>\n"
            "<STMTTRN>\n"
            "<TRNTYPE>DEBIT</TRNTYPE>\n"
            "<DTPOSTED>20260721000000[-3:BRT]</DTPOSTED>\n"
            "<TRNAMT>-91.56</TRNAMT>\n"
            "<FITID>fit-1</FITID>\n"
            "<MEMO>Zaffari Caxias</MEMO>\n"
            "</STMTTRN>\n"
            "<STMTTRN>\n"
            "<TRNTYPE>CREDIT</TRNTYPE>\n"
            "<DTPOSTED>20260709000000[-3:BRT]</DTPOSTED>\n"
            "<TRNAMT>705.91</TRNAMT>\n"
            "<FITID>fit-2</FITID>\n"
            "<MEMO>Pagamento recebido</MEMO>\n"
            "</STMTTRN>\n"
            "<STMTTRN>\n"
            "<TRNTYPE>CREDIT</TRNTYPE>\n"
            "<DTPOSTED>20260710000000[-3:BRT]</DTPOSTED>\n"
            "<TRNAMT>50.00</TRNAMT>\n"
            "<FITID>fit-3</FITID>\n"
            "<MEMO>Estorno</MEMO>\n"
            "</STMTTRN>\n"
            "</BANKTRANLIST>\n"
            "</OFX>\n"
        )

    def test_process_nubank_ofx(self):
        from unittest.mock import MagicMock
        headers = {"Authorization": "Bearer test-token", "userName": "test@test.com"}
        self.remote_repository.get_objects.return_value = []
        self.remote_repository.get_user.return_value = {"id": 1, "email": "test@test.com"}
        self.handler.analyze = MagicMock(return_value=([
            {'id': 'fit-1', 'category': 'Alimentação'},
            {'id': 'fit-3', 'category': 'Cashback'}
        ], 100))

        file_mock = MagicMock()
        file_mock.read.return_value = self._build_ofx().encode('latin-1')

        result = self.handler.process_nubank_ofx(file_mock, headers)

        self.assertEqual(result['status'], 'success')
        # "Pagamento recebido" is ignored, so 2 processed
        self.assertEqual(result['processed'], 2)
        self.assertEqual(result['inserted'], 2)
        self.assertEqual(self.remote_repository.insert.call_count, 2)

        first_tx = self.remote_repository.insert.call_args_list[0][1]["data"]
        self.assertEqual(first_tx["date"], "2026-07-21")
        self.assertEqual(first_tx["type"], "outcome")
        self.assertEqual(first_tx["value"], 91.56)
        self.assertEqual(first_tx["name"], "Zaffari Caxias")
        self.assertEqual(first_tx["app_name"], "Nubank da Suelen")
        self.assertEqual(first_tx["key"], "fit-1")
        self.assertEqual(first_tx["category"], "Alimentação")
        self.assertEqual(first_tx["is_installment"], "N")
        self.assertIn("Zaffari Caxias", first_tx["text"])
        self.assertEqual(first_tx["user_id"], 1)

        second_tx = self.remote_repository.insert.call_args_list[1][1]["data"]
        self.assertEqual(second_tx["type"], "income")
        self.assertEqual(second_tx["value"], 50.00)
        self.assertEqual(second_tx["key"], "fit-3")
        self.assertEqual(second_tx["category"], "Cashback")

    def test_process_nubank_ofx_duplicates(self):
        from unittest.mock import MagicMock
        headers = {"Authorization": "Bearer test-token", "userName": "test@test.com"}
        self.remote_repository.get_objects.return_value = [{"id": 123}]
        self.remote_repository.get_user.return_value = {"id": 1, "email": "test@test.com"}

        file_mock = MagicMock()
        file_mock.read.return_value = self._build_ofx().encode('latin-1')

        result = self.handler.process_nubank_ofx(file_mock, headers)

        self.assertEqual(result['status'], 'success')
        self.assertEqual(result['processed'], 2)
        self.assertEqual(result['inserted'], 0)
        self.remote_repository.insert.assert_not_called()


class TestAppRoutes(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        from unittest.mock import patch
        cls.starter_patch = patch('csctracker_py_core.starter.Starter.start')
        cls.starter_patch.start()
        import app
        cls.app_module = app
        cls.client = app.app.test_client()

    @classmethod
    def tearDownClass(cls):
        cls.starter_patch.stop()

    def test_process_dividends_endpoint_success(self):
        from unittest.mock import MagicMock
        original_handler = self.app_module.transaction_handler
        self.app_module.transaction_handler = MagicMock()
        self.app_module.transaction_handler.process_b3_dividends.return_value = {'status': 'success', 'processed': 1}
        
        try:
            res = self.client.post('/transactions/dividends', json={'file': 'SGVsbG8gd29ybGQ='})
            self.assertEqual(res.status_code, 201)
            self.assertEqual(res.get_json(), {'status': 'success', 'processed': 1})
            
            # Verify process_b3_dividends was called and got the decoded bytes
            call_arg = self.app_module.transaction_handler.process_b3_dividends.call_args[0][0]
            self.assertEqual(call_arg.read(), b'Hello world')
        finally:
            self.app_module.transaction_handler = original_handler

    def test_process_dividends_endpoint_success_with_data_uri(self):
        from unittest.mock import MagicMock
        original_handler = self.app_module.transaction_handler
        self.app_module.transaction_handler = MagicMock()
        self.app_module.transaction_handler.process_b3_dividends.return_value = {'status': 'success', 'processed': 1}
        
        try:
            res = self.client.post('/transactions/dividends', json={'file': 'data:application/octet-stream;base64,SGVsbG8gd29ybGQ='})
            self.assertEqual(res.status_code, 201)
            self.assertEqual(res.get_json(), {'status': 'success', 'processed': 1})
            
            # Verify process_b3_dividends was called and got the decoded bytes
            call_arg = self.app_module.transaction_handler.process_b3_dividends.call_args[0][0]
            self.assertEqual(call_arg.read(), b'Hello world')
        finally:
            self.app_module.transaction_handler = original_handler

    def test_process_dividends_endpoint_missing_file_field(self):
        res = self.client.post('/transactions/dividends', json={})
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.get_json()['status'], 'error')
        self.assertIn('No file in request body', res.get_json()['text'])

    def test_process_dividends_endpoint_empty_file_field(self):
        res = self.client.post('/transactions/dividends', json={'file': ''})
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.get_json()['status'], 'error')
        self.assertIn('File field is empty', res.get_json()['text'])

    def test_process_dividends_endpoint_invalid_base64(self):
        res = self.client.post('/transactions/dividends', json={'file': '!!!invalid!!!'})
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.get_json()['status'], 'error')
        self.assertIn('Invalid base64 format', res.get_json()['text'])

    def test_process_ofx_endpoint_success(self):
        from unittest.mock import MagicMock
        original_handler = self.app_module.transaction_handler
        self.app_module.transaction_handler = MagicMock()
        self.app_module.transaction_handler.process_nubank_ofx.return_value = {'status': 'success', 'processed': 1}

        try:
            res = self.client.post('/transactions/ofx', json={'file': 'SGVsbG8gd29ybGQ='})
            self.assertEqual(res.status_code, 201)
            self.assertEqual(res.get_json(), {'status': 'success', 'processed': 1})

            call_arg = self.app_module.transaction_handler.process_nubank_ofx.call_args[0][0]
            self.assertEqual(call_arg.read(), b'Hello world')
        finally:
            self.app_module.transaction_handler = original_handler

    def test_process_ofx_endpoint_missing_file_field(self):
        res = self.client.post('/transactions/ofx', json={})
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.get_json()['status'], 'error')
        self.assertIn('No file in request body', res.get_json()['text'])


if __name__ == "__main__":
    unittest.main()
