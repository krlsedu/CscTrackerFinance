import decimal
import io
import json
import logging
import openpyxl
import re
import uuid
from datetime import datetime

from csctracker_py_core.repository.http_repository import HttpRepository
from csctracker_py_core.repository.remote_repository import RemoteRepository
from csctracker_py_core.utils.request_info import RequestInfo
from csctracker_py_core.utils.utils import Utils
from dateutil.relativedelta import relativedelta


class Encoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, decimal.Decimal):
            return float(obj)


class TransactionHandler:
    def __init__(self, remote_repository: RemoteRepository, http_repository: HttpRepository):
        self.logger = logging.getLogger()
        self.remote_repository = remote_repository
        self.http_repository = http_repository
        pass

    def generate_transaction(self, json_info):
        try:
            text_ = json.loads(json_info['text'])
            info_ = text_['textInfo']
            if info_ == '':
                info_ = text_['textBig']
                if info_ == '':
                    info_ = text_['text']
                    if info_ == '':
                        info_ = text_['textSummary']

            title = text_['title']
            if title.find('NuPay') != -1:
                info_ = title + ' ' + info_
            self.transaction(info_, text_)
        except Exception as e:
            self.logger.info(json_info)
            self.logger.exception(e)

    def transaction(self, text_str, json_info):
        text_str = text_str.replace('  ', ' ')
        regex = r"(.*) ((.*\$)( | )(([1-9]\d{0,2}(.\d{3})*)|(([1-9]\d*)?\d))(\,\d\d)?)(.*)$"
        matches = re.finditer(regex, text_str)
        re_match = re.match(regex, text_str)
        if not re_match:
            regex = r"(.*) ((([1-9]\d{0,2}(.\d{3})*)|(([1-9]\d*)?\d))(\,\d\d)?)(.*)$"
            re_match = re.match(regex, text_str)
            if not re_match:
                regex = r"(.*) ((.*\$)( | |  )(([1-9]\d{0,2}(.\d{3})*)|(([1-9]\d*)?\d))(\,\d\d)?)(.*)$"
                re_match = re.match(regex, text_str)
        if re_match:
            for matchNum, match in enumerate(matches, start=1):
                transaction = {}
                type = match.group(1)
                value = match.group(2)
                status = match.group(11)
                package = json_info['packageName']
                app_name = json_info['appName']
                f = int(json_info['postTime'])
                if "devolvemos o iof" in text_str.lower():
                    transaction['type'] = "income"
                    transaction['category'] = "Cashback"
                    transaction['name'] = f"IOF {app_name}"
                else:
                    transaction['type'] = self.get_type(type, status)
                    try:
                        transaction['name'] = self.get_name(text_str)
                    except:
                        transaction['name'] = self.get_name(status)

                transaction['value'] = self.get_value(value)

                transaction['package_name'] = package
                transaction['app_name'] = app_name
                transaction['text'] = text_str
                try:
                    key_ = json_info['key']
                    key_ = re.sub('[^A-Za-z0-9]+', '_', key_)
                    transaction['key'] = key_
                except:
                    pass
                fromtimestamp = datetime.fromtimestamp(f / 1000)
                transaction['date'] = fromtimestamp \
                    .strftime('%Y-%m-%d %H:%M:%S')
                installments_ = self.get_installments(text_str)
                if installments_ > 1:
                    self.split_transaction(installments_, transaction)
                else:
                    transaction['is_installment'] = 'N'
                    self.save_transaction(transaction)

    def split_transaction(self, installments_, transaction):
        value = transaction['value'] / installments_
        value = round(value, 2)
        transaction['value'] = value
        transaction['is_installment'] = 'S'
        transaction['installment_id'] = str(uuid.uuid4())
        date_str_ = transaction['date']
        text_ = transaction['text']
        for i in range(installments_):
            transaction['text'] = text_ + f" {i + 1}/{installments_}"
            try:
                date_ = datetime.strptime(date_str_, '%Y-%m-%d %H:%M:%S')
            except:
                date_ = datetime.strptime(date_str_, '%Y-%m-%d')
            if i > 0 and ('id' in transaction and transaction['id'] is not None):
                del transaction['id']
            date_ += relativedelta(months=+i)
            transaction['date'] = date_.strftime('%Y-%m-%d')
            if i == 0:
                self.save_transaction(transaction, num_parcs=installments_)
            else:
                self.save_transaction(transaction, num_parcs=0)

    def get_type(self, text, status):
        regex = r"(.*)(compra|Compra|Recebemos.*pagamento)(.*)$"
        if re.match(regex, text):
            regex = r"(.*)(não.*autorizada)(.*)$"
            if re.match(regex, status):
                return 'unknown'
            regex = r"(.*)(estornada|cancelada)(.*)$"
            if re.match(regex, status):
                return "income"
            return "outcome"
        else:
            regex = r"(.*)(recebeu.*Pix|Recebemos.*PIX|Recebemos.*transferência|recebeu.*transferência)(.*)$"
            if re.match(regex, text):
                return "income"
        return "unknown"

    def get_value(self, text):
        regex = r"((([1-9]\d{0,2}(.\d{3})*)|(([1-9]\d*)?\d))(\,\d\d))$"
        matches = re.finditer(regex, text)

        for matchNum, match in enumerate(matches, start=1):
            value = match.group(1)
            return float(value.strip().replace('.', '').replace(',', '.'))
        return float(0)

    def get_name(self, text):
        regex = r"(?:em\s)(.*?)(?:\sàs)(.*)$"
        match = re.search(regex, text)
        if match:
            return match.group(1).strip()
        else:
            regex = r"(.*)(em )(.*)( para o cartão)(.*)$"
            match = re.search(regex, text)
            if match:
                return match.group(3).strip()
            else:
                regex = r"(em\s)(.*?)(\scom\s|\sfoi)"
                match = re.search(regex, text)
                if match:
                    return match.group(2).strip()
                else:
                    regex = r"(.*)(em )(.*)(\.)$"
                    if re.match(regex, text):
                        return re.match(regex, text).group(3).strip()
                    else:
                        regex = r"(.*)(de )(.*)(\.)$"
                        if re.match(regex, text):
                            return re.match(regex, text).group(3).strip()
        return "unknown"

    def get_installments(self, text):
        regex = r"\((\d+x)\)"
        match = re.search(regex, text)
        if match:
            number = re.search(r"\d+", match.group(1))
            return int(number.group()) if number else 1
        else:
            return 1

    def save_transactions(self, transactions, headers):
        for transaction in transactions:
            installments_ = self.get_installments(transaction['text'])
            if 'is_installment' not in transaction:
                transaction['is_installment'] = 'N'
            if installments_ > 1 and transaction['is_installment'] == 'S':
                filter_ = {
                    'key': transaction['key'],
                    'is_installment': 'S',
                    'installment_id': transaction['installment_id']
                }
                transactions_ = self.remote_repository.get_objects("transactions", data=filter_, headers=headers)
                for transaction_ in transactions_:
                    transaction_['category'] = transaction['category']
                    transaction_['name'] = transaction['name']
                    self.remote_repository.insert("transactions",
                                                  data=transaction_,
                                                  headers=headers)
                    suffix = f" 1/{installments_}"
                    if transaction_['text'].endswith(suffix):
                        self.check_and_save_cashback(transaction_, num_parcs=installments_, headers=headers)
            elif installments_ > 1 and (
                    transaction['is_installment'] == 'N'
                    or 'id' not in transaction
                    or transaction['id'] is None
            ):
                self.split_transaction(installments_, transaction)
            else:
                if 'id' not in transaction or transaction['id'] is None:
                    transaction['is_installment'] = 'N'
                self.save_transaction(transaction)
        return "OK"

    def save_transaction(self, transaction, num_parcs=1):
        try:
            headers = self.http_repository.get_headers()
            is_new = 'id' not in transaction or transaction['id'] is None
            if is_new:
                try:
                    if 'key' in transaction and transaction['key'] is not None and transaction['key'].startswith('ia_extractor_'):
                        exists = self.remote_repository.get_objects("transactions",
                                                                    keys=["app_name", "value", "date", "type"],
                                                                    data=transaction,
                                                                    headers=headers)
                        if exists.__len__() == 0:
                            exists = self.remote_repository.get_objects("transactions",
                                                                    keys=["key", "value", "date"],
                                                                    data=transaction,
                                                                    headers=headers)
                        if exists.__len__() > 0:
                            transaction['category'] = 'Ignored'
                        pass
                    else:
                        exists = self.remote_repository.get_objects("transactions",
                                                                    keys=["key", "value", "date"],
                                                                    data=transaction,
                                                                    headers=headers)
                        self.logger.info(exists)
                        if exists.__len__() > 0:
                            self.logger.info(f"Transaction already saved-> {transaction['key']} -> {transaction}")
                            Utils.inform_to_client(transaction, "urgent",
                                                   headers,
                                                   f"Transaction already saved-> {transaction['key']} "
                                                   f"- {transaction['value']} - {transaction['date']}")
                            transaction['category'] = 'Ignored'
                except Exception as e:
                    self.logger.exception(e)
            response = self.remote_repository.insert("transactions",
                                                     data=transaction,
                                                     headers=headers)

            self.check_and_save_cashback(transaction, num_parcs, headers)

            return response, 201
        except Exception as e:
            self.logger.exception(e)
            return {
                'text': 'transaction not saved',
                'status': 'error',
                'error': str(e)
            }, 400

    def check_and_save_cashback(self, transaction, num_parcs=1, headers=None):
        if headers is None:
            headers = self.http_repository.get_headers()
        cashback_exists = False
        if 'key' in transaction and transaction['key'] is not None:
            try:
                cashback_key = f"{transaction['key']}_cashback"
                exists_cashback = self.remote_repository.get_objects("transactions",
                                                                     data={'key': cashback_key},
                                                                     headers=headers)
                if exists_cashback:
                    for cb in exists_cashback:
                        if cb.get('category') != 'Ignored':
                            cashback_exists = True
                            break
            except Exception as e:
                self.logger.exception(e)

        if not cashback_exists and transaction.get('app_name') == 'Nubank' and transaction.get('type') == 'outcome' and transaction.get('category') != 'Ignored' and num_parcs > 0:
            try:
                cashback_transaction = transaction.copy()
                cashback_transaction.pop('id', None)
                cashback_transaction.pop('is_installment', None)
                cashback_transaction.pop('installment_id', None)
                cashback_transaction['type'] = 'income'
                cashback_transaction['name'] = 'Nubank'
                cashback_transaction['category'] = 'Cashback'
                
                original_value = transaction.get('value', 0)
                cashback_value = round((original_value * num_parcs) * 0.0125, 2)
                cashback_transaction['value'] = cashback_value
                
                original_text = transaction.get('text', '')
                cashback_transaction['text'] = f"Cashback de {cashback_value} referente a {original_text}"
                
                if 'key' in cashback_transaction and cashback_transaction['key'] is not None:
                    cashback_transaction['key'] = f"{cashback_transaction['key']}_cashback"
                    
                self.save_transaction(cashback_transaction)
            except Exception as e:
                self.logger.exception(e)

    def process_b3_dividends(self, file_storage, headers):
        file_bytes = file_storage.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        max_r = sheet.max_row
        
        processed_count = 0
        inserted_count = 0
        try:
            request_id = RequestInfo.get_correlation_id()
        except Exception:
            request_id = None
        if not request_id:
            request_id = f"CscTrackerBff-{uuid.uuid4()}"
            
        user_id = 1
        if headers:
            try:
                user = self.remote_repository.get_user(headers)
                if user and 'id' in user:
                    user_id = user['id']
            except Exception as e:
                self.logger.warning(f"Could not retrieve user from headers, using default user_id = 1. Error: {e}")
        
        for r in range(2, max_r + 1):
            produto = sheet.cell(row=r, column=1).value
            pagamento = sheet.cell(row=r, column=2).value
            tipo_evento = sheet.cell(row=r, column=3).value
            instituicao = sheet.cell(row=r, column=4).value
            quantidade = sheet.cell(row=r, column=5).value
            preco_unitario = sheet.cell(row=r, column=6).value
            valor_liquido = sheet.cell(row=r, column=7).value
            
            if produto is None:
                continue
            produto_str = str(produto).strip()
            if not produto_str or produto_str == "" or produto_str.lower() == "total":
                continue
                
            if '-' in produto_str:
                name = produto_str.split('-')[0].strip()
            else:
                name = produto_str
                
            date_formatted = self._format_date(pagamento)
            if not date_formatted:
                continue
                
            qty_str = self._format_quantity(quantidade)
            pu_str = self._to_currency_str(preco_unitario)
            vl_str = self._to_currency_str(valor_liquido)
            vl_float = self._to_float(valor_liquido)
            
            tipo_evento_str = str(tipo_evento).strip() if tipo_evento is not None else ""
            instituicao_str = str(instituicao).strip() if instituicao is not None else ""
            
            text = f"{tipo_evento_str} recebido, referente a {qty_str} cotas de {produto_str} no valor de {pu_str} por cota, total: {vl_str} na instituição {instituicao_str}"
            
            if "NU INVESTIMENTOS S.A. - CTVM" in instituicao_str:
                app_name = "Nubank"
            elif "BANCO BTG PACTUAL S/A." in instituicao_str:
                app_name = "BTG"
            else:
                app_name = instituicao_str
                
            qty_int = self._to_int(quantidade)
            pu_float = self._to_float(preco_unitario)
            
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')
            
            dividend_data = {
                'ticker': name,
                'data_pagamento': date_formatted,
                'tipo_evento': tipo_evento_str,
                'quantidade': qty_int,
                'preco_unitario': pu_float,
                'user_id': user_id,
                'request_id': request_id,
                'last_update': now_str,
                'created_at': now_str
            }
            
            exists = self.remote_repository.get_objects(
                "dividends_b3",
                keys=["ticker", "data_pagamento", "tipo_evento", "quantidade", "preco_unitario", "user_id"],
                data=dividend_data,
                headers=headers
            )
            
            processed_count += 1
            if len(exists) == 0:
                key_uuid = str(uuid.uuid4())
                
                transaction = {
                    'date': date_formatted,
                    'type': 'income',
                    'value': vl_float,
                    'name': name,
                    'package_name': None,
                    'app_name': app_name,
                    'text': text,
                    'user_id': user_id,
                    'last_update': now_str,
                    'category': 'Proventos',
                    'key': key_uuid,
                    'copy': None,
                    'request_id': request_id,
                    'is_installment': 'N',
                    'installment_id': None
                }
                
                self.remote_repository.insert("transactions", data=transaction, headers=headers)
                self.remote_repository.insert("dividends_b3", data=dividend_data, headers=headers)
                
                inserted_count += 1
                
        return {
            'status': 'success',
            'processed': processed_count,
            'inserted': inserted_count
        }

    def _to_float(self, val):
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        if isinstance(val, str):
            val = val.strip()
            clean = val.replace("R$", "").replace(" ", "")
            if ',' in clean:
                clean = clean.replace(".", "").replace(",", ".")
            try:
                return float(clean)
            except ValueError:
                return 0.0
        return 0.0

    def _to_int(self, val):
        if val is None:
            return 0
        if isinstance(val, (int, float)):
            return int(val)
        if isinstance(val, str):
            val = val.strip()
            try:
                return int(float(val.replace(',', '.')))
            except ValueError:
                return 0
        return 0

    def _to_currency_str(self, val):
        if val is None:
            return "R$ 0,00"
        if isinstance(val, (int, float)):
            parts = f"{val:.2f}".split('.')
            integer_part = parts[0]
            decimal_part = parts[1]
            
            reversed_integer = integer_part[::-1]
            groups = [reversed_integer[i:i+3] for i in range(0, len(reversed_integer), 3)]
            formatted_integer = '.'.join(groups)[::-1]
            
            return f"R$ {formatted_integer},{decimal_part}"
        if isinstance(val, str):
            val = val.strip()
            if not val.startswith("R$"):
                f_val = self._to_float(val)
                return self._to_currency_str(f_val)
            return val
        return str(val)

    def _format_quantity(self, val):
        if val is None:
            return "0"
        if isinstance(val, (int, float)):
            if val == int(val):
                return str(int(val))
            return str(val)
        if isinstance(val, str):
            val = val.strip()
            try:
                val_float = float(val.replace(',', '.'))
                if val_float == int(val_float):
                    return str(int(val_float))
                return str(val_float)
            except ValueError:
                return val
        return str(val)

    def _format_date(self, val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.strftime('%Y-%m-%d')
        if hasattr(val, 'strftime'):
            return val.strftime('%Y-%m-%d')
        if isinstance(val, str):
            val = val.strip()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d'):
                try:
                    dt = datetime.strptime(val, fmt)
                    return dt.strftime('%Y-%m-%d')
                except ValueError:
                    continue
        return str(val)
